import openpyxl


wb = openpyxl.load_workbook('censuspopdata.xlsx')
#sheet = wb.active # cho phep hoat dong tren file excel do
#sheet.title = 'Sheet1' #doi trang tinh
#print(sheet.title)

#lay phan tu cua 1 hang dat truoc
"""
sheet = wb['Population by Census Tract']
arr = []
for i in range(1,sheet.max_column+1):
    arr.append(str(sheet.cell(row=14,column=i).value))
print(arr)
"""
#lay phan tu cua 1 hang dat truoc
"""
sheet = wb['Population by Census Tract']
for i in list(sheet.rows)[14]:
    print(i.value)
"""